"""错题导出：Anki 卡包（genanki, MIT）与 Excel 表格（openpyxl, MIT）。

可选依赖，未安装时给出提示：
    pip install genanki    # Anki .apkg 导出
    pip install openpyxl   # Excel .xlsx 导出
"""
import os
from typing import Any


def export_anki(items: list[dict[str, Any]], path: str) -> str:
    """把错题导出为 Anki 卡包（.apkg）。

    genanki（MIT 许可证, https://github.com/kerrickstaley/genanki）
    生成标准 Anki 卡组，卡片正面为题号+题干，背面为答案+解析。
    """
    try:
        import genanki
    except ImportError:
        raise RuntimeError("导出 Anki 需要安装 genanki: pip install genanki")

    model = genanki.Model(
        1712341001,  # 固定 ID，避免每次导出生成新模型
        "考研错题卡",
        fields=[
            {"name": "Subject"},
            {"name": "Question"},
            {"name": "Answer"},
            {"name": "Analysis"},
        ],
        templates=[
            {
                "name": "错题复习",
                "qfmt": "{{Subject}}<br><hr>{{Question}}",
                "afmt": '{{FrontSide}}<hr id="answer">'
                        "<b>答案:</b> {{Answer}}<br><br>"
                        "<b>解析:</b> {{Analysis}}",
            },
        ],
    )
    deck = genanki.Deck(1712341002, "考研错题本")

    for it in items:
        deck.add_note(genanki.Note(
            model=model,
            fields=[
                it.get("subject") or "未分类",
                it.get("question", "") or "",
                it.get("correct_answer") or it.get("my_answer") or "",
                it.get("analysis", "") or "",
            ],
        ))

    genanki.Package(deck).write_to_file(path)
    return path


def export_excel(items: list[dict[str, Any]], path: str) -> str:
    """把错题导出为 Excel 表格（.xlsx）。

    openpyxl（MIT 许可证, https://github.com/openpyxl/openpyxl）
    适合用 Excel/WPS 做二次筛选与打印。
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("导出 Excel 需要安装 openpyxl: pip install openpyxl")

    wb = Workbook()
    ws = wb.active
    ws.title = "错题本"

    headers = ["ID", "科目", "来源", "题干", "我的答案", "正确答案",
               "解析", "标签", "错误次数", "创建时间", "复盘时间"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DFF0FF")
        cell.alignment = Alignment(vertical="center")

    for it in items:
        ws.append([
            it.get("id"),
            it.get("subject") or "",
            it.get("source") or "",
            it.get("question") or "",
            it.get("my_answer") or "",
            it.get("correct_answer") or "",
            it.get("analysis") or "",
            "、".join(it.get("tags") or []),
            it.get("wrong_count", 1),
            it.get("created_at") or "",
            it.get("reviewed_at") or "",
        ])

    # 宽度与自动换行优化（使用 get_column_letter 支持任意列数）
    widths = [6, 10, 14, 50, 16, 16, 40, 12, 10, 18, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    os.makedirs(os.path.dirname(os.path.abspath(path)) or ".", exist_ok=True)
    wb.save(path)
    return path